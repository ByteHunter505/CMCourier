"""Shared pytest fixtures.

Houses session-scoped fixtures that prepare deterministic test data
(generated from Python sources) so binary blobs stay out of git history.
"""

from __future__ import annotations

from pathlib import Path

import pytest

_FIXTURES_DIR = Path(__file__).parent / "fixtures" / "sources"


@pytest.fixture(scope="session", autouse=True)
def _generate_xlsx_fixtures() -> None:
    """Generate XLSX fixture files at session start if they do not exist.

    The data is hardcoded so the resulting files are deterministic, and
    regeneration is sub-second. Files are gitignored so binary diffs do not
    pollute the repo. Imports happen inside the fixture so collection is
    cheap when these tests are not selected.
    """
    _FIXTURES_DIR.mkdir(parents=True, exist_ok=True)

    sample_xlsx = _FIXTURES_DIR / "sample.xlsx"
    if not sample_xlsx.exists():
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("openpyxl Workbook returned no active sheet")
        # Same shape as sample.csv: header + 5 rows. One row with blank Age
        # (for NaN normalization test) and JUANPEREZ01 appearing twice (for
        # get_by_fields equality test).
        ws.append(["Name", "Age", "Birth"])
        ws.append(["JUANPEREZ01", "30", "1995-06-12"])
        ws.append(["MARIAGOMEZ02", None, "1988-11-23"])
        ws.append(["JUANPEREZ01", "45", "1980-04-05"])
        ws.append(["PEPELOPEZ03", "25", "2000-01-15"])
        ws.append(["TESTUSER04", None, "2010-07-30"])
        wb.save(sample_xlsx)

    multi_sheet_xlsx = _FIXTURES_DIR / "multi_sheet.xlsx"
    if not multi_sheet_xlsx.exists():
        from openpyxl import Workbook

        wb = Workbook()
        ws1 = wb.active
        if ws1 is None:
            raise RuntimeError("openpyxl Workbook returned no active sheet")
        ws1.title = "Sheet1"
        ws1.append(["Col"])
        ws1.append(["sheet1_value"])
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["Col"])
        ws2.append(["sheet2_value_a"])
        ws2.append(["sheet2_value_b"])
        wb.save(multi_sheet_xlsx)
